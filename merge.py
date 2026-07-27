"""Сборка папки однотипных таблиц в один сводный отчёт."""

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUPPORTED = {".xlsx", ".xlsm", ".csv"}
SOURCE_COLUMN = "Файл-источник"

HEADER_FILL = PatternFill("solid", fgColor="2F3640")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D5D8DC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class MergeError(Exception):
    """Ошибка, которую можно показать пользователю без стектрейса."""


def collect_files(folder: Path, recursive: bool) -> list[Path]:
    if not folder.exists():
        raise MergeError(f"Папка не найдена: {folder}")
    if not folder.is_dir():
        raise MergeError(f"Это не папка: {folder}")

    pattern = "**/*" if recursive else "*"
    files = [
        path
        for path in sorted(folder.glob(pattern))
        if path.is_file()
        and path.suffix.lower() in SUPPORTED
        and not path.name.startswith("~$")
    ]
    if not files:
        raise MergeError(
            f"В папке {folder} нет таблиц. Поддерживаются: {', '.join(sorted(SUPPORTED))}"
        )
    return files


def coerce(value: str):
    """CSV отдаёт всё строками — числа возвращаем числами, иначе итоги не посчитать."""
    text = value.strip()
    if not text:
        return None

    normalized = text.replace(" ", "").replace(" ", "").replace(",", ".")
    digits = normalized.lstrip("-")

    # Ведущий ноль — это артикул или телефон, а не число: 007 должно остаться 007.
    if len(digits) > 1 and digits.startswith("0") and not digits.startswith("0."):
        return text

    if re.fullmatch(r"-?\d+", normalized):
        return int(normalized)
    if re.fullmatch(r"-?\d+\.\d+", normalized):
        return float(normalized)
    return text


def read_csv(path: Path) -> tuple[list[str], list[list]]:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            with path.open(encoding=encoding, newline="") as handle:
                sample = handle.read(8192)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                rows = [row for row in csv.reader(handle, dialect) if any(cell.strip() for cell in row)]
            break
        except UnicodeDecodeError:
            continue
    else:
        raise MergeError(f"Не удалось определить кодировку: {path.name}")

    if not rows:
        return [], []
    return [cell.strip() for cell in rows[0]], [[coerce(cell) for cell in row] for row in rows[1:]]


def read_excel(path: Path, sheet: str | None) -> tuple[list[str], list[list]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise MergeError(f"В файле {path.name} нет листа «{sheet}»")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = [
            list(row)
            for row in worksheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
    finally:
        workbook.close()

    if not rows:
        return [], []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return header, rows[1:]


def read_table(path: Path, sheet: str | None) -> tuple[list[str], list[list]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_excel(path, sheet)


def build_columns(tables: list[tuple[Path, list[str], list[list]]]) -> list[str]:
    """Порядок колонок первого файла, недостающие из остальных — в конец."""
    columns: list[str] = []
    for _, header, _ in tables:
        for name in header:
            if name and name not in columns:
                columns.append(name)
    return columns


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def autosize(worksheet, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 3, 11), 52)


def write_data_sheet(worksheet, columns: list[str], rows: list[list]) -> dict[str, float]:
    worksheet.append(columns)
    widths = {index: len(name) for index, name in enumerate(columns, start=1)}

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 28

    totals: dict[str, float] = {}
    numeric_hits: dict[str, int] = {}

    for row in rows:
        worksheet.append(row)
        for index, (name, value) in enumerate(zip(columns, row), start=1):
            if value is None:
                continue
            widths[index] = max(widths.get(index, 0), len(str(value)))
            if is_number(value):
                totals[name] = totals.get(name, 0) + value
                numeric_hits[name] = numeric_hits.get(name, 0) + 1

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = CELL_BORDER
            if is_number(cell.value):
                cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

    autosize(worksheet, widths)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    # Колонка считается числовой, только если числа в большинстве строк.
    return {
        name: (value, numeric_hits[name])
        for name, value in totals.items()
        if numeric_hits.get(name, 0) >= max(1, len(rows) // 2)
    }


def write_summary_sheet(
    worksheet,
    per_file: list[tuple[str, int]],
    totals: dict[str, tuple[float, int]],
    total_rows: int,
) -> None:
    worksheet.append(["Источник", "Строк"])
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")

    widths = {1: len("Источник"), 2: len("Строк")}
    for name, count in per_file:
        worksheet.append([name, count])
        widths[1] = max(widths[1], len(name))

    worksheet.append(["Всего", total_rows])
    for cell in worksheet[worksheet.max_row]:
        cell.font = TOTAL_FONT

    if totals:
        worksheet.append([])
        worksheet.append(["Числовая колонка", "Сумма", "Среднее"])
        for cell in worksheet[worksheet.max_row]:
            cell.font = TOTAL_FONT
        widths[3] = len("Среднее")
        for name, (total, count) in totals.items():
            worksheet.append([name, round(total, 2), round(total / count, 2)])
            widths[1] = max(widths[1], len(name))

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if is_number(cell.value):
                cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

    autosize(worksheet, widths)
    worksheet.freeze_panes = "A2"


def merge(folder: Path, output: Path, sheet: str | None, recursive: bool) -> tuple[int, int]:
    files = collect_files(folder, recursive)
    print(f"Найдено таблиц: {len(files)}")

    tables: list[tuple[Path, list[str], list[list]]] = []
    for path in files:
        try:
            header, rows = read_table(path, sheet)
        except MergeError:
            raise
        except Exception as error:
            print(f"  ! {path.name} — пропущен ({type(error).__name__}: {error})")
            continue

        if not header:
            print(f"  ! {path.name} — пропущен (файл пустой)")
            continue

        tables.append((path, header, rows))
        print(f"  + {path.name} — строк: {len(rows)}")

    if not tables:
        raise MergeError("Ни один файл не удалось прочитать")

    columns = build_columns(tables) + [SOURCE_COLUMN]

    merged: list[list] = []
    per_file: list[tuple[str, int]] = []
    for path, header, rows in tables:
        index_by_name = {name: index for index, name in enumerate(header) if name}
        for row in rows:
            ordered = []
            for name in columns[:-1]:
                index = index_by_name.get(name)
                ordered.append(row[index] if index is not None and index < len(row) else None)
            ordered.append(path.name)
            merged.append(ordered)
        per_file.append((path.name, len(rows)))

    workbook = Workbook()
    totals = write_data_sheet(workbook.active, columns, merged)
    workbook.active.title = "Данные"
    write_summary_sheet(workbook.create_sheet("Сводка"), per_file, totals, len(merged))

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return len(tables), len(merged)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Собирает папку однотипных таблиц в один сводный Excel-отчёт.",
        epilog="Пример: python3 merge.py ./отчёты -o итог.xlsx",
    )
    parser.add_argument("folder", type=Path, help="папка с таблицами (.xlsx, .xlsm, .csv)")
    parser.add_argument("-o", "--out", type=Path, default=Path("сводный-отчёт.xlsx"), help="куда сохранить результат")
    parser.add_argument("-s", "--sheet", help="имя листа в исходных файлах (по умолчанию первый)")
    parser.add_argument("-r", "--recursive", action="store_true", help="искать таблицы и во вложенных папках")
    args = parser.parse_args()

    try:
        files_count, rows_count = merge(args.folder, args.out, args.sheet, args.recursive)
    except MergeError as error:
        sys.stdout.flush()  # иначе в пайпе ошибка обгонит прогресс
        print(f"\nОшибка: {error}", file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        sys.stdout.flush()
        print("\nПрервано пользователем", file=sys.stderr)
        return 130

    print(f"\nГотово: {files_count} файлов, {rows_count} строк → {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
